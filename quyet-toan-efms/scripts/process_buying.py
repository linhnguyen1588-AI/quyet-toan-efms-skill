import os
import sys
import openpyxl
import glob
from datetime import datetime
from copy import copy
import re

sys.stdout.reconfigure(encoding='utf-8')

def get_suffix(s):
    # Lấy chuỗi số ở đuôi (ít nhất 4 chữ số) để làm matching linh hoạt
    m = re.search(r'\d{4,}$', s)
    if m:
        return m.group(0)
    return None

def process():
    print("1. Đọc dữ liệu từ BANG KE SOTRAN FORTUNE T5.2026.xlsx...")
    wb_bk = openpyxl.load_workbook(r"D:\SOTRANS LINH 2\LINH\BUYING VT\THANG 5.2026\BANG KE SOTRAN FORTUNE T5.2026.xlsx", data_only=True)
    sheet_bk = wb_bk.active
    
    bills = []
    for r in range(11, sheet_bk.max_row + 1):
        val_bill = sheet_bk.cell(row=r, column=4).value # Cột D (Số Bill)
        val_cont = sheet_bk.cell(row=r, column=5).value # Cột E (Số Cont)
        
        if val_bill is not None and str(val_bill).strip() != "":
            bill_str = str(val_bill).strip().upper()
            has_cont = False
            cont_str = ""
            if val_cont is not None and str(val_cont).strip() != "":
                cont_str = str(val_cont).strip().upper()
                if "HÀNG LẺ" not in cont_str and "HANG LE" not in cont_str:
                    has_cont = True
            bills.append((bill_str, cont_str, has_cont))
            
    print(f"-> Đã lấy {len(bills)} số bill từ bảng kê.")

    print("\n2. Đọc dữ liệu từ các file SAN LUONG VT (T1 đến T5)...")
    sl_files = glob.glob(r"D:\SOTRANS LINH 2\LINH\BUYING VT\SAN LUONG VT*.xlsx")
    
    bill_to_amount = {}
    suffix_to_amount = {}
    cont_to_amount = {}
    
    for filepath in sl_files:
        if not os.path.exists(filepath):
            continue
        wb_sl = openpyxl.load_workbook(filepath, data_only=True)
        sheet_sl = wb_sl.active
        
        for r in range(20, sheet_sl.max_row + 1):
            client = sheet_sl.cell(row=r, column=3).value
            if client and "FORTUNE" in str(client).upper():
                waybill = sheet_sl.cell(row=r, column=2).value
                booking = sheet_sl.cell(row=r, column=4).value
                sl_cont = sheet_sl.cell(row=r, column=6).value # Cột F (Số Cont)
                ak_val = sheet_sl.cell(row=r, column=37).value # Cột AK
                
                w_str = str(waybill).strip().upper() if waybill is not None else ""
                b_str = str(booking).strip().upper() if booking is not None else ""
                c_str = str(sl_cont).strip().upper() if sl_cont is not None else ""
                
                # Cập nhật mapping Exact và Suffix
                if w_str:
                    bill_to_amount[w_str] = ak_val
                    w_suf = get_suffix(w_str)
                    if w_suf: suffix_to_amount[w_suf] = ak_val
                if b_str:
                    bill_to_amount[b_str] = ak_val
                    b_suf = get_suffix(b_str)
                    if b_suf: suffix_to_amount[b_suf] = ak_val
                if c_str:
                    cont_to_amount[c_str] = ak_val
                    
    print(f"-> Đã map được {len(bill_to_amount)} bills (waybill/booking) chính xác.")
    print(f"-> Đã tạo mapping phụ dựa trên số đuôi ({len(suffix_to_amount)} mẫu).")
    print(f"-> Đã tạo mapping dự phòng dựa trên số cont ({len(cont_to_amount)} mẫu).")

    print("\n3. Cập nhật file Buying FOTUN T5. VT.xlsx...")
    buying_path = r"D:\SOTRANS LINH 2\LINH\BUYING VT\THANG 5.2026\Buying FOTUN T5. VT.xlsx"
    
    # Chuẩn bị wb_done
    wb_done = openpyxl.load_workbook(buying_path)
    sheet_done = wb_done["Sheet2"]
    
    # Lưu lại template (dòng 2)
    template_row = []
    for c in range(2, 19): # Cột B đến R
        template_row.append(sheet_done.cell(row=2, column=c).value)
        
    # Xóa dòng cũ trong wb_done
    if sheet_done.max_row >= 2:
        sheet_done.delete_rows(2, sheet_done.max_row - 1)
        
    # Chuẩn bị wb_check cho các bill lỗi
    wb_check = openpyxl.load_workbook(buying_path)
    sheet_check = wb_check["Sheet2"]
    if sheet_check.max_row >= 2:
        sheet_check.delete_rows(2, sheet_check.max_row - 1)
        
    today_str = datetime.now().strftime('%d/%m/%Y')
    
    row_done = 2
    row_check = 2
    
    for idx, (bill, cont_str, has_cont) in enumerate(bills):
        amount = 0
        if has_cont:
            if bill in bill_to_amount:
                amount = bill_to_amount[bill]
            else:
                suf = get_suffix(bill)
                if suf and suf in suffix_to_amount:
                    amount = suffix_to_amount[suf]
            
            # Backup: Quét lại bằng số cont nếu bill không có dữ liệu
            if not amount and cont_str and cont_str in cont_to_amount:
                amount = cont_to_amount[cont_str]
        
        if amount:
            # Có data -> Ghi vào DONE
            sheet_done.cell(row=row_done, column=1, value=bill)
            for c_idx, val in enumerate(template_row):
                sheet_done.cell(row=row_done, column=c_idx + 2, value=val)
            sheet_done.cell(row=row_done, column=9, value=amount) # Cột I
            sheet_done.cell(row=row_done, column=19, value=row_done - 1) # Cột S (Note)
            row_done += 1
        else:
            # Không có data (hoặc không có số cont) -> Ghi vào CHECK
            sheet_check.cell(row=row_check, column=1, value=bill)
            for c_idx, val in enumerate(template_row):
                sheet_check.cell(row=row_check, column=c_idx + 2, value=val)
            sheet_check.cell(row=row_check, column=9, value=0) # Cột I
            
            # Đổi ngày thực hiện cho cột M (13) và P (16)
            sheet_check.cell(row=row_check, column=13, value=today_str)
            sheet_check.cell(row=row_check, column=16, value=today_str)
            
            sheet_check.cell(row=row_check, column=19, value=row_check - 1) # Cột S
            row_check += 1
            
    output_done = r"D:\SOTRANS LINH 2\LINH\BUYING VT\THANG 5.2026\Buying FOTUN T5. VT_DONE.xlsx"
    output_check = r"D:\SOTRANS LINH 2\LINH\BUYING VT\THANG 5.2026\Buying FOTUN T5. VT_CHECK.xlsx"
    
    wb_done.save(output_done)
    if row_check > 2:
        wb_check.save(output_check)
        
    print(f"\n=> Đã xử lý xong!")
    print(f"=> Đã lưu file DONE: {row_done - 2} bills có data.")
    if row_check > 2:
        print(f"=> Đã lưu file CHECK: {row_check - 2} bills KHÔNG có data (hoặc không có số cont) (Đã cập nhật ngày thực hiện).")

if __name__ == "__main__":
    process()
