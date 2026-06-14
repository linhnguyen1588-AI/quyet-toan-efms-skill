import openpyxl

data = {
    "TRHU6112075": {"price": 1700000, "inv": "00024825", "date": "13/05/2026"},
    "CCLU7677844": {"price": 1650000, "inv": "00053616", "date": "14/05/2026"},
    "CSNU5510286": {"price": 1650000, "inv": "00053618", "date": "14/05/2026"},
    "HMMU4284158": {"price": 1650000, "inv": "3254", "date": "15/05/2026"},
    "HMMU6016133": {"price": 1650000, "inv": "3256", "date": "15/05/2026"},
    "FITU5519190": {"price": 601852, "inv": "129567", "date": "16/05/2026"},
    "HMMU6852576": {"price": 1650000, "inv": "3342", "date": "18/05/2026"},
    "TXGU4122556": {"price": 1650000, "inv": "3346", "date": "18/05/2026"},
    "BSIU9671780": {"price": 1650000, "inv": "00003841", "date": "27/05/2026"}
}

file_path = r"C:\Users\LEGION\OneDrive\Máy tính\TOAN LUC T5 NLP\BUYING TOAN LUC PARITAS.xlsx"
print(f"Opening {file_path}...")
wb = openpyxl.load_workbook(file_path)
ws = wb.active

updated_count = 0
for row in range(2, ws.max_row + 1):
    cont = ws.cell(row=row, column=2).value
    if cont and isinstance(cont, str):
        cont = cont.strip()
        if cont in data:
            # Column I (9) = Unit Price
            ws.cell(row=row, column=9).value = data[cont]["price"]
            # Column O (15) = Invoice No
            ws.cell(row=row, column=15).value = data[cont]["inv"]
            # Column P (16) = Invoice Date
            ws.cell(row=row, column=16).value = data[cont]["date"]
            updated_count += 1

wb.save(file_path)
print(f"Done! Đã cập nhật thành công {updated_count} dòng.")
