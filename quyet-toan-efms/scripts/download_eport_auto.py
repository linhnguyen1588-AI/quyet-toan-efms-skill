import asyncio
import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
from playwright.async_api import async_playwright

# Force UTF-8 encoding
sys.stdout.reconfigure(encoding='utf-8')

async def wait_if_captcha(page):
    while await page.locator("text=Vui lòng nhập mã này").count() > 0 or await page.locator("text=What code is in the image").count() > 0:
        print("\n[!] PHÁT HIỆN CAPTCHA! Cảnh sát ePort đang chặn đường.")
        print("[!] Sếp VUI LÒNG GIẢI MÃ CAPTCHA trên màn hình ePort nhé! Robot đang đứng chờ...")
        await page.wait_for_timeout(5000)

async def main():
    excel_path = r"D:\workspace-ai\HOA DON\HAOHUA\XUAT THANG 6 HOA DON EPORT.xlsx"
    out_dir = r"D:\workspace-ai\HOA DON\HAOHUA"
    
    if not os.path.exists(out_dir):
        os.makedirs(out_dir)
        
    print("Đang đọc file Excel...")
    try:
        df = pd.read_excel(excel_path)
        bookings = df['BOOKING'].dropna().astype(str).unique().tolist()
        bookings = [b for b in bookings if b.strip() and b.strip() != 'nan']
        top5 = bookings[:5]
    except Exception as e:
        print("Lỗi đọc/lưu file Excel:", e)
        return

    if not top5:
        print("Không tìm thấy số Book nào trong file!")
        return

    print(f"Sẽ tiến hành quét 5 Book: {top5}")
    
    failed_books = []

    async with async_playwright() as p:
        # Giảm thời gian chờ xuống, tắt headless=False để thấy web múa
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(accept_downloads=True)
        page = await context.new_page()
        
        print("\n==================================================")
        print("Trình duyệt đã mở. ĐANG TỰ ĐỘNG ĐĂNG NHẬP VÀ CHẠY SIÊU TỐC...")
        print("==================================================\n")
        
        await page.goto("https://eport.saigonnewport.com.vn/")
        
        # Tự động điền tài khoản và đăng nhập
        try:
            await wait_if_captcha(page)
            await page.wait_for_timeout(3000)
            tax_loc = page.locator("input[type='text']:visible").first
            await tax_loc.fill("0314436809")
            pass_loc = page.locator("input[type='password']:visible").first
            await pass_loc.fill("Sotrans1234@")
            login_btn = page.locator("button:visible, input[type='submit']:visible").filter(has_text="Đăng nhập").first
            await login_btn.click()
            print("Đã bấm đăng nhập...")
            await page.wait_for_timeout(3000)
            await wait_if_captcha(page)
        except Exception as e:
            print("Lỗi đăng nhập tự động:", e)
        
        one_month_ago = (datetime.now() - timedelta(days=30)).strftime("%d/%m/%Y")
        
        for book in top5:
            print(f"\n>>> Đang xử lý Book: {book}")
            try:
                await page.goto("https://eport.saigonnewport.com.vn/FullContainerDelivery")
                await wait_if_captcha(page)
                # Chờ input Ghi chú xuất hiện
                await page.wait_for_selector("input, textarea", timeout=10000)
                
                # Nhập "Từ ngày"
                inputs = await page.locator("input, textarea").all()
                for inp in inputs:
                    ph = await inp.get_attribute("placeholder") or ""
                    name = await inp.get_attribute("name") or ""
                    if "từ ngày" in ph.lower() or "fromdate" in name.lower():
                        await inp.fill("")
                        await inp.fill(one_month_ago)
                        
                # Nhập Ghi chú
                found_note = False
                for inp in inputs:
                    ph = await inp.get_attribute("placeholder") or ""
                    name = await inp.get_attribute("name") or ""
                    if "ghi ch" in ph.lower() or "note" in name.lower() or "remark" in name.lower():
                        await inp.fill(book)
                        found_note = True
                        break
                        
                if not found_note:
                    await page.evaluate(f'''() => {{
                        let inputs = document.querySelectorAll('input, textarea');
                        for (let el of inputs) {{
                            let html = el.outerHTML.toLowerCase();
                            if (html.includes('ghichu') || html.includes('note') || html.includes('remark')) {{
                                el.value = '{book}';
                                el.dispatchEvent(new Event('input', {{ bubbles: true }}));
                                break;
                            }}
                        }}
                    }}''')
                    
                # Bấm Tìm kiếm
                btns = await page.locator("button, a").all()
                for btn in btns:
                    text = await btn.text_content() or ""
                    if "tìm kiếm" in text.lower() or "lấy dữ liệu" in text.lower():
                        await btn.click()
                        break
                        
                print("  -> Đã bấm tìm kiếm, đợi kết quả...")
                await page.wait_for_timeout(3000)
                
                # Bấm Tải Hóa Đơn (Đã fix CHỈ TÌM TRONG BẢNG GRID)
                # Dùng .dx-datagrid hoặc table để khoanh vùng
                grid = page.locator(".dx-datagrid, table").first
                
                inv_locators = [
                    "a[title*='Hóa đơn']", 
                    "a[title*='hóa đơn']", 
                    "i.fa-file-invoice", 
                    "i.fa-download"
                ]
                
                downloaded = False
                for loc in inv_locators:
                    elements = await grid.locator(loc).all()
                    if elements:
                        print(f"  -> Đã tìm thấy {len(elements)} nút hóa đơn trong bảng. Bấm tải siêu tốc...")
                        for idx, el in enumerate(elements):
                            try:
                                async with page.expect_download(timeout=10000) as download_info:
                                    # Click chuột nhanh, bypass mọi thẻ chắn ngang
                                    await el.click(force=True)
                                    download = await download_info.value
                                    
                                    base_name = book
                                    file_path = os.path.join(out_dir, f"{base_name}.pdf")
                                    counter = 1
                                    while os.path.exists(file_path):
                                        file_path = os.path.join(out_dir, f"{base_name}-0{counter}.pdf")
                                        counter += 1
                                        
                                    await download.save_as(file_path)
                                    print(f"  [v] TẢI XONG: {file_path}")
                                    downloaded = True
                            except Exception as e:
                                pass
                        
                        if downloaded:
                            break 
                        
                if not downloaded:
                    print(f"  [!] Book {book} không có dữ liệu hóa đơn. Đang lưu mã nguồn HTML để phân tích lỗi...")
                    try:
                        content = await page.content()
                        with open(f"eport_dom_{book}.html", "w", encoding="utf-8") as f:
                            f.write(content)
                    except Exception as e:
                        print("  [!] Lỗi khi lưu HTML:", e)
                    failed_books.append(book)
            except Exception as e:
                print(f"  [!] Lỗi khi xử lý Book {book}")
                failed_books.append(book)
                
        print("\n🎉 ĐÃ CHẠY XONG 5 BOOK!")
        await browser.close()
        
    # Xử lý tô màu vàng cho các book lỗi
    if failed_books:
        print(f"\n[+] Đang tô vàng {len(failed_books)} Book lỗi...")
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            
            # Tìm cột BOOKING
            booking_col_idx = None
            for row in ws.iter_rows(min_row=1, max_row=5):
                for cell in row:
                    if cell.value and "BOOKING" in str(cell.value).upper():
                        booking_col_idx = cell.column
                        break
                if booking_col_idx: break
            
            if booking_col_idx:
                yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
                for row in ws.iter_rows(min_row=2):
                    cell_val = str(row[booking_col_idx - 1].value).strip() if row[booking_col_idx - 1].value else ""
                    if cell_val in failed_books:
                        for cell in row:
                            cell.fill = yellow_fill
                
                try:
                    wb.save(excel_path)
                    print(f"  [v] Đã tô vàng xong!")
                except PermissionError:
                    print(f"  [!] LỖI: Vui lòng tắt file Excel rồi chạy lại để tô màu.")
            else:
                print("  [!] Không tìm thấy cột BOOKING.")
        except Exception as e:
            pass
    else:
        print("\n[+] Tuyệt vời! Tất cả đều tải thành công 100%.")

if __name__ == "__main__":
    asyncio.run(main())
