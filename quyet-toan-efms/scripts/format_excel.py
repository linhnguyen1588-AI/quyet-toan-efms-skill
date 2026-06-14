import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

file_path = r"D:\SOTRANS LINH 2\LINH\BUYING VT\BẢNG KÊ TÁI XUẤT GN-KHỐI CƯỚC.xlsx"

try:
    wb = openpyxl.load_workbook(file_path)
    
    # Styles
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Dark Blue
    header_font = Font(color="FFFFFF", bold=True, size=11, name="Calibri")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    data_font = Font(size=11, name="Calibri")
    data_alignment_left = Alignment(horizontal="left", vertical="center")
    data_alignment_center = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Find the header row (assume it's the first non-empty row or just row 1)
        # Let's just assume row 1 is the header for now, but sometimes there are titles.
        # So we look for the row with the most columns.
        max_cols = ws.max_column
        max_rows = ws.max_row
        
        # Determine header row by finding the first row with multiple string values
        header_row_idx = 1
        for row in range(1, min(10, max_rows + 1)):
            count = sum(1 for cell in ws[row] if cell.value is not None)
            if count > 1:
                header_row_idx = row
                break
                
        # Freeze panes below header
        ws.freeze_panes = ws.cell(row=header_row_idx + 1, column=1)
        
        # Apply header styles
        for col in range(1, max_cols + 1):
            cell = ws.cell(row=header_row_idx, column=col)
            if cell.value is not None:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border
        
        # Apply data styles
        for row in range(header_row_idx + 1, max_rows + 1):
            for col in range(1, max_cols + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None or ws.cell(row=header_row_idx, column=col).value is not None:
                    cell.font = data_font
                    cell.border = thin_border
                    # Align numbers to right, text to left, but center might look good.
                    # Let's keep it simple: general alignment, but vertically centered.
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0'
                    # cell.alignment = Alignment(vertical="center")
                    
        # Auto-fit column widths
        for col_idx in range(1, max_cols + 1):
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            max_length = 0
            for row in range(1, max_rows + 1):
                cell = ws.cell(row=row, column=col_idx)
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            # Add some padding
            adjusted_width = (max_length + 2)
            # Cap width at 50 to prevent crazy wide columns
            if adjusted_width > 50:
                adjusted_width = 50
            if adjusted_width < 10:
                adjusted_width = 10
                
            ws.column_dimensions[col_letter].width = adjusted_width

    output_path = r"D:\SOTRANS LINH 2\LINH\BUYING VT\BẢNG KÊ TÁI XUẤT GN-KHỐI CƯỚC_Formatted.xlsx"
    wb.save(output_path)
    print(f"Success! Saved to {output_path}")

except Exception as e:
    import traceback
    traceback.print_exc()
