import os
import re
import asyncio
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from playwright.async_api import async_playwright
import sys
import time

sys.stdout.reconfigure(encoding='utf-8', line_buffering=True)

YELLOW_FILL = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")

def get_declaration_numbers(file_path):
    try:
        df = pd.read_excel(file_path, header=None)
    except Exception as e:
        print(f"Lỗi khi đọc file Excel: {e}")
        return [], None, None

    target_col = None
    target_row = None
    for row in range(len(df)):
        for col in range(len(df.columns)):
            val = str(df.iloc[row, col]).strip().upper()
            if val == 'SỐ TỜ KHAI':
                target_col = col
                target_row = row
                break
        if target_col is not None:
            break

    if target_col is None:
        print("Không tìm thấy cột 'SỐ TỜ KHAI'")
        return [], None, None

    declarations = []
    tb_col = None
    for col in range(len(df.columns)):
        if str(df.iloc[target_row, col]).strip().upper() == 'SỐ TB PHÍ':
            tb_col = col
            if col == 5:
                break
    
    # Create PDF directory if not exists
    pdf_dir = os.path.join(os.path.dirname(file_path), "PDFs")
    if not os.path.exists(pdf_dir):
        os.makedirs(pdf_dir)

    for row in range(target_row + 1, len(df)):
        val = str(df.iloc[row, target_col]).strip()
        if val and val.lower() != 'nan':
            parts = re.split(r'[\n,]', val)
            for p in parts:
                p = p.strip()
                if p:
                    declarations.append((row, p))

    return declarations, target_col, tb_col

def write_to_excel_incremental(file_path, declarations_map, tb_map, missing_map, tb_col):
    try:
        output_file = file_path.replace(".xlsx", "_Done.xlsx")
        if not os.path.exists(output_file):
            wb = openpyxl.load_workbook(file_path)
        else:
            wb = openpyxl.load_workbook(output_file)
            
        ws = wb.active
        
        row_updates = {}
        missing_rows = set()
        
        for row, tk in declarations_map:
            if tk in tb_map:
                if row not in row_updates:
                    row_updates[row] = []
                row_updates[row].append(tb_map[tk])
            if tk in missing_map:
                missing_rows.add(row)
                
        for row, tb_list in row_updates.items():
            excel_row = row + 1
            excel_col = tb_col + 1
            
            existing_val = ws.cell(row=excel_row, column=excel_col).value
            existing_val = str(existing_val).strip() if existing_val else ""
            
            all_tb = [p.strip() for p in existing_val.split('\n') if p.strip() and p.strip().lower() != 'nan']
            all_tb.extend(tb_list)
            unique_tb = list(dict.fromkeys(all_tb))
            
            ws.cell(row=excel_row, column=excel_col).value = "\n".join(unique_tb)
            
        # Tô màu vàng cho các dòng không có thông báo phí
        for row in missing_rows:
            excel_row = row + 1
            for col in range(1, ws.max_column + 1):
                ws.cell(row=excel_row, column=col).fill = YELLOW_FILL
            
        wb.save(output_file)
    except Exception as e:
        print(f"Lỗi khi ghi Excel: {e}")

async def run_scraper():
    file_path = r"D:\workspace-ai\CSHT\CSHT HAOHUA\BANG KE CSHT HAOHUA T5.xlsx"
    declarations, tk_col, tb_col = get_declaration_numbers(file_path)
    
    unique_decls = list(dict.fromkeys([d[1] for d in declarations]))
    print(f"-> Tổng cộng {len(unique_decls)} mã tờ khai.")
    
    if not unique_decls:
        return

    save_dir = r"D:\workspace-ai\CSHT\CSHT HAOHUA\PDFs"
    os.makedirs(save_dir, exist_ok=True)
    
    # Check what is already done
    processed_tk = set()
    output_file = file_path.replace(".xlsx", "_Done.xlsx")
    if os.path.exists(output_file):
        try:
            wb = openpyxl.load_workbook(output_file, data_only=True)
            ws = wb.active
            for row in range(2, ws.max_row + 1):
                tk_val = str(ws.cell(row=row, column=tk_col+1).value).strip()
                tb_val = ws.cell(row=row, column=tb_col+1).value
                bg_color = ws.cell(row=row, column=1).fill.start_color.index if ws.cell(row=row, column=1).fill else None
                
                if bg_color == 'FFFFFF00':
                    parts = re.split(r'[\n,]', tk_val)
                    for p in parts:
                        p = p.strip()
                        if p:
                            processed_tk.add(p)
                elif tb_val and str(tb_val).strip() != 'None' and str(tb_val).strip() != '':
                    parts = re.split(r'[\n,]', tk_val)
                    for p in parts:
                        p = p.strip()
                        if p:
                            if str(tb_val).strip() == '000000000000':
                                processed_tk.add(p)
                            else:
                                pdf_path = os.path.join(save_dir, f"THONG BÁO PHI CSHT _ {p}.pdf")
                                if os.path.exists(pdf_path):
                                    processed_tk.add(p)
                                else:
                                    print(f"-> Tờ khai {p} có số TB trong Excel nhưng THIẾU file PDF, sẽ tải lại!")
        except Exception as e:
            print(f"Lỗi khi đọc file Done: {e}")

    async with async_playwright() as p:
        browser = await p.chromium.launch(headless=False)
        context = await browser.new_context(accept_downloads=True)
        page = await context.new_page()

        await page.goto("https://thuphihatang.tphcm.gov.vn/dang-nhap")
        
        try:
            await page.fill('input[type="text"]', "0315428529")
            await page.fill('input[type="password"]', "SSotrans123@1")
        except:
            pass

        print("\n" + "="*70)
        print("VUI LÒNG THAO TÁC TRÊN TRÌNH DUYỆT:")
        print("1. Nhập CAPTCHA và bấm Đăng nhập.")
        print("2. Vào trang Tra cứu biên lai.")
        print("="*70 + "\n")

        while True:
            try:
                if await page.query_selector('input[name="SO_TK"]'):
                    break
            except:
                pass
            await asyncio.sleep(1)

        print("-> Đã phát hiện giao diện tra cứu.")
        print("-> Dừng lại 15 giây để sếp chỉnh lại Ngày (ví dụ: Từ 01/05/2026 đến hôm nay).")
        for i in range(15, 0, -1):
            print(f"Tool sẽ tự động chạy sau {i} giây...")
            await asyncio.sleep(1)
            
        print("-> Hết giờ chờ, bắt đầu tự động quét các tờ khai!")

        tb_map = {}
        missing_map = set()

        for idx, tk in enumerate(unique_decls):
            if tk in processed_tk:
                print(f"[{idx+1}/{len(unique_decls)}] Bỏ qua tờ khai đã xử lý: {tk}")
                continue
                
            print(f"[{idx+1}/{len(unique_decls)}] Đang xử lý tờ khai: {tk}")
            
            # Check login session
            if "dang-nhap" in page.url:
                print("  -> Phiên đăng nhập đã hết hạn! Vui lòng đăng nhập lại và chọn Ngày.")
                while True:
                    try:
                        if await page.query_selector('input[name="SO_TK"]'):
                            break
                    except:
                        pass
                    await asyncio.sleep(1)
                print("  -> Đã đăng nhập lại thành công. Tiếp tục quét...")
            
            # Đóng các popup cũ nếu có
            try:
                close_btns = await page.query_selector_all('.jconfirm-closeIcon:visible, button.close:visible, button:has-text("Đóng"):visible')
                for btn in close_btns:
                    await btn.click(timeout=1000, force=True)
            except:
                pass
            
            try:
                # Clear the table body to prevent reading old data
                await page.evaluate("document.querySelector('#TBLDANHSACH tbody').innerHTML = '<tr id=\"my-loading-flag\"><td colspan=\"15\">Đang tải...</td></tr>'")
                
                # Nhập SO_TK và Search
                await page.fill('input[name="SO_TK"]', tk)
                await page.click('.btnSearch', force=True)
                
                # Wait up to 30 seconds for the website to replace the table content
                await page.wait_for_function("!document.getElementById('my-loading-flag')", timeout=30000)
                await asyncio.sleep(1) # Thêm 1s để DOM render ổn định
            except Exception as e:
                print(f"  -> Lỗi khi click tìm kiếm hoặc bảng không phản hồi: {e}")
                await page.keyboard.press("Escape")
                print("  -> Bỏ qua tờ khai này để thử lại sau.")
                # We intentionally DO NOT add it to missing_map, so it won't be saved as YELLOW.
                # It will be retried in the next run.
                continue
                
            try:
                popups = await page.query_selector_all('.jconfirm-closeIcon')
                for popup in popups:
                    await popup.click()
            except:
                pass

            try:
                rows = await page.query_selector_all('#TBLDANHSACH tbody tr')
                if rows and len(rows) > 0:
                    tds = await rows[0].query_selector_all('td')
                    if len(tds) > 10:
                        # Thử lấy thông báo trước
                        btn_lay_tb = await rows[0].query_selector('.btnXemThongBaoNP')
                        if btn_lay_tb:
                            print("  -> Bấm Lấy thông báo...")
                            await btn_lay_tb.click(force=True)
                            await asyncio.sleep(5) # Đợi hệ thống sinh thông báo
                            
                            # Tắt popup báo thành công
                            try:
                                popups = await page.query_selector_all('.jconfirm-closeIcon')
                                for p in popups: await p.click()
                            except:
                                await page.keyboard.press("Escape")
                                
                            # Re-read
                            rows = await page.query_selector_all('#TBLDANHSACH tbody tr')
                            tds = await rows[0].query_selector_all('td')

                        so_tb = await tds[10].inner_text()
                        so_tb = so_tb.strip().split('\n')[0].strip() # Clean msgId
                        
                        actual_tk = await tds[5].inner_text()
                        actual_tk = actual_tk.strip()
                        
                        if actual_tk != tk:
                            print(f"  -> Cảnh báo: Tờ khai trên web ({actual_tk}) không khớp với tờ khai đang tìm ({tk})!")
                        
                        if so_tb and so_tb != "0" and so_tb != "000000000000":
                            tb_map[tk] = so_tb
                            print(f"  -> Số TB Phí: {so_tb}")
                            
                            # Tìm nút Tải thông báo nộp phí
                            try:
                                dl_btn = None
                                # Dùng query_selector_all để lấy tất cả các nút Tải, sau đó chọn cái đang hiển thị
                                btns = await page.query_selector_all('.btnPrintThongBaoNP, button:has-text("Tải thông báo nộp phí"), button:has-text("In thông báo")')
                                for b in btns:
                                    if await b.is_visible():
                                        dl_btn = b
                                        break
                                
                                if dl_btn:
                                    print("  -> Đã tìm thấy nút Tải, đang tiến hành tải PDF...")
                                    try:
                                        async with page.expect_download(timeout=30000) as download_info:
                                            await dl_btn.click(force=True)
                                        download = await download_info.value
                                        pdf_path = os.path.join(save_dir, f"THONG BÁO PHI CSHT _ {tk}.pdf")
                                        await download.save_as(pdf_path)
                                        print(f"  -> Đã tải PDF thành công: {pdf_path}")
                                    except Exception as e:
                                        print(f"  -> Lỗi khi tải PDF (có thể là do mạng hoặc quá thời gian): {e}")
                                        
                                    # Click nút Đóng sau khi tải xong để trở về trang cũ
                                    try:
                                        close_btns = page.locator('.jconfirm-closeIcon:visible, button.close:visible, button:has-text("Đóng"):visible')
                                        count = await close_btns.count()
                                        for i in range(count):
                                            try:
                                                await close_btns.nth(i).click(timeout=1000)
                                            except: pass
                                        await asyncio.sleep(1)
                                    except:
                                        pass
                                else:
                                    print("  -> Không tìm thấy nút Tải thông báo.")
                            except Exception as e:
                                print(f"  -> Lỗi khi click nút tải: {e}")
                        else:
                            print(f"  -> Chưa có thông báo phí (000000000000).")
                            missing_map.add(tk)
                    else:
                        print("  -> Không tìm thấy dữ liệu.")
                        missing_map.add(tk)
                else:
                    print("  -> Không tìm thấy dữ liệu.")
                    missing_map.add(tk)
            except Exception as e:
                print(f"  -> Lỗi khi đọc bảng dữ liệu: {e}")
                # We do NOT add to missing_map if there is an error. We want it to retry later!
            
            # Save to Excel every 10 items to improve speed
            if tb_col is not None and (idx + 1) % 10 == 0:
                if tb_map or missing_map:
                    write_to_excel_incremental(file_path, declarations, tb_map, missing_map, tb_col)
                    processed_tk.update(tb_map.keys())
                    processed_tk.update(missing_map)
                    tb_map.clear()
                    missing_map.clear()

        # Final save for any remaining items
        if tb_col is not None and (tb_map or missing_map):
            write_to_excel_incremental(file_path, declarations, tb_map, missing_map, tb_col)
            tb_map.clear()
            missing_map.clear()

        await browser.close()
        print("-> Hoàn tất toàn bộ quy trình!")

if __name__ == "__main__":
    asyncio.run(run_scraper())
