import os
import json
import argparse
import urllib.request
import urllib.parse
import datetime
import difflib
import pandas as pd
import sys
import re
import unicodedata
import math

def remove_accents(input_str):
    nfkd_form = unicodedata.normalize('NFKD', input_str)
    res = "".join([c for c in nfkd_form if not unicodedata.combining(c)])
    return res.replace('đ', 'd').replace('Đ', 'D')

# Reconfigure stdout/stderr to use UTF-8 encoding
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')
if hasattr(sys.stderr, 'reconfigure'):
    sys.stderr.reconfigure(encoding='utf-8')

# Resolve paths relative to this script
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SKILL_DIR = os.path.dirname(SCRIPT_DIR)
CONFIG_FILE = os.path.join(SKILL_DIR, "config.json")
MAP_FILE = os.path.join(SKILL_DIR, "company_map.json")

def load_json(filepath):
    if os.path.exists(filepath):
        with open(filepath, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def fetch_api_data(api_url, company, month, year):
    params = {"month": month, "year": year, "sheet": company}
    query_str = urllib.parse.urlencode(params)
    url = f"{api_url}?{query_str}"
    
    print(f"[INFO] Fetching data from API: {url}")
    req = urllib.request.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    try:
        with urllib.request.urlopen(req) as response:
            data = json.loads(response.read().decode('utf-8'))
            return data
    except Exception as e:
        print(f"[ERROR] API Request failed: {e}")
        return None

def fetch_local_excel_data(file_path, month=None, year=None, company=None, bill_col=None, cont_col=None, tk_col=None):
    print(f"[INFO] Fetching data from local Excel: {file_path}")
    shipments = []
    try:
        try:
            xls = pd.ExcelFile(file_path)
            sheet_names = xls.sheet_names
        except Exception as e:
            print(f"[WARNING] pd.ExcelFile failed ({e}). Retrying with engine='calamine'...")
            xls = pd.ExcelFile(file_path, engine='calamine')
            sheet_names = xls.sheet_names
            
        for sheet_name in sheet_names:
            try:
                df = pd.read_excel(xls, sheet_name=sheet_name, header=None)
            except Exception:
                df = pd.read_excel(xls, sheet_name=sheet_name, header=None, engine='calamine')
                
            block_parsing_mode = False
            in_month_block = False
            
            if not month or (company and "haohua" in company.lower()):
                block_parsing_mode = True
                in_month_block = True
                
            for index, row in df.iterrows():
                try:
                    col_b = str(row[1]).strip().upper() if pd.notna(row[1]) else ""
                    val = col_b
                    
                    if not block_parsing_mode:
                        if str(month) in col_b and str(year) in col_b:
                            block_parsing_mode = True
                            in_month_block = True
                            print(f"[DEBUG] Found month header in sheet '{sheet_name}': {val}")
                            continue
                    else:
                        if not in_month_block:
                            continue
                            
                        if "THÁNG" in val or "MONTH" in val or "-" in val:
                            import re
                            if re.search(r'\d{4}-\d{2}', val) or re.search(r'\d{1,2}/\d{4}', val) or "THÁNG" in val:
                                if str(month) not in val or str(year) not in val:
                                    in_month_block = False
                                    continue
                                    
                    if not in_month_block:
                        continue
                        
                    if company and "haohua" in company.lower():
                        # Haohua specific format
                        bill_index = bill_col if bill_col is not None else 1
                        so_bill = str(row[bill_index]).strip() if pd.notna(row[bill_index]) else "" # Default Cột B
                        if not so_bill or so_bill.lower() == 'nan' or len(so_bill) < 4:
                            continue
                            
                        # Khai báo hải quan
                        tk_index = tk_col if tk_col is not None else 2
                        declarations_str = str(row[tk_index]).strip() if pd.notna(row[tk_index]) else ""
                        if not declarations_str or declarations_str.lower() == 'nan':
                            continue
                            
                        # Luồng ở cột H (row[7]) hoặc E (row[4])
                        phan_luong_1 = str(row[4]).strip().lower() if pd.notna(row[4]) else ""
                        phan_luong_2 = str(row[7]).strip().lower() if pd.notna(row[7]) else ""
                        phan_luong = phan_luong_2 if phan_luong_2 and phan_luong_2 != 'nan' else phan_luong_1
                        if not phan_luong or phan_luong == 'nan':
                            phan_luong = "xanh"
                            
                        try:
                            c_idx = cont_col if cont_col is not None else 5
                            cont_total = float(str(row[c_idx]).replace(',', '')) if pd.notna(row[c_idx]) and str(row[c_idx]).strip() != '' else 1
                        except:
                            cont_total = 1
                            
                        loai_cont = str(row[6]).strip().upper() if pd.notna(row[6]) else ""
                        if "40" in loai_cont:
                            cont_40 = cont_total
                            cont_20 = 0
                        else:
                            cont_20 = cont_total
                            cont_40 = 0
                            
                    else:
                        so_bill = str(row[2]).strip() if pd.notna(row[2]) else "" # Cột C
                        if not so_bill or so_bill.lower() == 'nan' or len(so_bill) < 4:
                            continue
                            
                        try:
                            cont_20 = float(str(row[5]).replace(',', '')) if pd.notna(row[5]) and str(row[5]).strip() != '' else 0
                        except:
                            cont_20 = 0
                        try:
                            cont_40 = float(str(row[6]).replace(',', '')) if pd.notna(row[6]) and str(row[6]).strip() != '' else 0
                        except:
                            cont_40 = 0
                            
                        default_luong = "xanh"
                        if company:
                            c_lower = company.lower()
                            if "paritas" in c_lower or "toàn lực" in c_lower or "toan luc" in c_lower:
                                default_luong = "vàng"
                                
                        try:
                            phan_luong = str(row[4]).strip().lower() if pd.notna(row[4]) and str(row[4]).strip().lower() != 'nan' else default_luong
                        except Exception:
                            phan_luong = default_luong
                            
                        declarations_str = str(row[17]).strip() # Cột R
                        if not declarations_str or declarations_str.lower() == 'nan':
                            continue # Bỏ qua vì không có số tờ khai

                    is_special = False
                    if " " in so_bill:
                        is_special = True
                        
                    raw_bill = so_bill
                    so_bill = so_bill.replace(" ", "")
                    
                    total_cont = cont_20 + cont_40
                    if total_cont == 0:
                        total_cont = 1
                        
                    declarations = [d.strip() for d in declarations_str.replace('\n', ',').split(',') if d.strip()]
                    if not declarations:
                        continue
                        
                    to_khai_count = len([d for d in declarations if d])
                    if to_khai_count == 0:
                        to_khai_count = 1
                        
                    phu_thu_hq = 0
                    try:
                        val_o = str(row[14]).strip() if pd.notna(row[14]) else ""
                        if val_o and val_o.lower() != 'nan':
                            phu_thu_hq = float(val_o.replace(',', ''))
                    except ValueError:
                        pass

                    phu_thu_bx = 0
                    try:
                        val_p = str(row[15]).strip() if pd.notna(row[15]) else ""
                        if val_p and val_p.lower() != 'nan':
                            phu_thu_bx = float(val_p.replace(',', ''))
                    except ValueError:
                        pass
                        
                    note = str(row[20]).strip() if pd.notna(row[20]) and str(row[20]).strip().lower() != 'nan' else ""

                    to_khai_joined = ", ".join(declarations)
                    if getattr(args, 'group_tk', True):  # Default to True
                        shipment = {
                            "so_bill": so_bill,
                            "raw_bill": raw_bill,
                            "is_special": is_special,
                            "phan_luong": phan_luong,
                            "cont_20": cont_20,
                            "cont_40": cont_40,
                            "total_cont": total_cont,
                            "to_khai_count": to_khai_count,
                            "to_khai": to_khai_joined,
                            "declarations": declarations,
                            "note": note,
                            "phu_thu_hq": phu_thu_hq,
                            "phu_thu_bx": phu_thu_bx
                        }
                        shipments.append(shipment)
                    else:
                        for decl in declarations:
                            shipment = {
                                "so_bill": so_bill,
                                "raw_bill": raw_bill,
                                "is_special": is_special,
                                "phan_luong": phan_luong,
                                "cont_20": cont_20,
                                "cont_40": cont_40,
                                "total_cont": total_cont,
                                "to_khai_count": 1, 
                                "to_khai": decl,
                                "declarations": [decl] if decl else [],
                                "note": note,
                                "phu_thu_hq": phu_thu_hq,
                                "phu_thu_bx": phu_thu_bx
                            }
                            shipments.append(shipment)
                except Exception as row_e:
                    print(f"[DEBUG] Skipping row due to error: {row_e}")
                    continue
        
        return shipments
    except Exception as e:
        print(f"[ERROR] Local Excel parsing failed: {e}")
        return None

def extract_tariff_matrix(tariff_file):
    print(f"[INFO] Parsing Tariff file: {tariff_file}")
    try:
        # Load all sheets to find the one with "BẢNG KHOÁN CHI PHÍ LÀM HÀNG"
        try:
            xls = pd.ExcelFile(tariff_file)
            sheet_names = xls.sheet_names
        except Exception:
            xls = pd.ExcelFile(tariff_file, engine='calamine')
            sheet_names = xls.sheet_names
            
        target_sheet = None
        for sheet_name in sheet_names:
            try:
                df_temp = pd.read_excel(tariff_file, sheet_name=sheet_name, header=None)
            except Exception:
                df_temp = pd.read_excel(tariff_file, sheet_name=sheet_name, header=None, engine='calamine')
                
            # Flatten the first 10 rows to search for the string
            head_str = df_temp.head(10).astype(str).to_string().upper()
            if "BẢNG KHOÁN CHI PHÍ LÀM HÀNG" in head_str or "KHOÁN TTHQ" in head_str:
                target_sheet = sheet_name
                break
                
        if not target_sheet:
            print("[ERROR] Could not find the appropriate sheet in Tariff file.")
            return None
            
        try:
            df = pd.read_excel(tariff_file, sheet_name=target_sheet, header=None)
        except Exception:
            df = pd.read_excel(tariff_file, sheet_name=target_sheet, header=None, engine='calamine')
        
        # We will extract prices for Xanh, Vàng, Đỏ for 1..6 cont.
        tariff_matrix = {
            "xanh": {},
            "vàng": {},
            "đỏ": {}
        }
        
        for idx, row in df.iterrows():
            row_text = " ".join([str(x).strip().lower() for x in row if pd.notna(x)])
            if "khoán tthq" in row_text:
                if "xanh" in row_text:
                    key = "xanh"
                elif "vàng" in row_text or "vang" in row_text:
                    key = "vàng"
                elif "đỏ" in row_text or "do" in row_text:
                    key = "đỏ"
                else:
                    continue
                
                # Extract numeric values from the row (excluding the row title)
                values = [int(float(x)) for x in row[2:] if pd.notna(x) and str(x).replace('.','',1).isdigit()]
                
                # Assuming the values map to 1, 2, 3, 4, 5, 6 containers
                # Sometimes they are duplicated for 20' and 40'. We just need unique values or every second value.
                # In the analyzed file, columns were (2,3) for 1 cont, (4,5) for 2 cont, etc.
                # So we can just take values[0], values[2], values[4], values[6], values[8], values[10]
                if len(values) >= 12:
                    tariff_matrix[key][1] = values[0]
                    tariff_matrix[key][2] = values[2]
                    tariff_matrix[key][3] = values[4]
                    tariff_matrix[key][4] = values[6]
                    tariff_matrix[key][5] = values[8]
                    tariff_matrix[key][6] = values[10]
                elif len(values) >= 6:
                    # If they are not duplicated
                    tariff_matrix[key][1] = values[0]
                    tariff_matrix[key][2] = values[1]
                    tariff_matrix[key][3] = values[2]
                    tariff_matrix[key][4] = values[3]
                    tariff_matrix[key][5] = values[4]
                    tariff_matrix[key][6] = values[5]
                else:
                    print(f"[WARNING] Unrecognized format for {key} row in Tariff. Found values: {values}")
                    
        return tariff_matrix

    except Exception as e:
        print(f"[ERROR] Failed to extract tariff matrix: {e}")
        return None

def get_tariff_price(tariff_matrix, phan_luong, total_cont):
    phan_luong = str(phan_luong).strip().lower()
    luong_key = None
    if "xanh" in phan_luong:
        luong_key = "xanh"
    elif "vàng" in phan_luong or "vang" in phan_luong:
        luong_key = "vàng"
    elif "đỏ" in phan_luong or "do" in phan_luong:
        luong_key = "đỏ"
        
    if not luong_key or luong_key not in tariff_matrix or not tariff_matrix[luong_key]:
        # Fallback default if not found
        print(f"[WARNING] Could not find tariff for luồng '{phan_luong}'. Using 380000 default.")
        return 380000
        
    cont_index = total_cont
    if cont_index > 6:
        cont_index = 6
    if cont_index < 1:
        cont_index = 1
        
    return tariff_matrix[luong_key].get(cont_index, 380000)

def main():
    parser = argparse.ArgumentParser(description="EFMS Buying Builder")
    parser.add_argument("--company", type=str, required=True, help="Company name (e.g. Paritas)")
    parser.add_argument("--month", type=int, required=True, help="Billing month")
    parser.add_argument("--year", type=int, required=True, help="Billing year")
    parser.add_argument('--limit', type=int, help="Limit output rows")
    parser.add_argument('--input-file', type=str, help="Path to local Excel file (for debugging)")
    parser.add_argument("--split", type=int, required=False, default=1, help="Number of files to split the output into")
    parser.add_argument("--bill-col", type=int, help="Zero-based index of Bill column")
    parser.add_argument("--cont-col", type=int, help="Zero-based index of Container column")
    parser.add_argument("--tk-col", type=int, help="Zero-based index of To Khai column")
    parser.add_argument("--group-tk", action="store_true", help="Group multiple To Khai into one shipment (default)")
    parser.add_argument("--no-group-tk", action="store_false", dest="group_tk", help="Do not group To Khai")
    parser.set_defaults(group_tk=True)
    global args
    args = parser.parse_args()

    config = load_json(CONFIG_FILE)
    company_map = load_json(MAP_FILE)
    
    base_output_dir = config.get("output_dir", "./")
    api_url = config.get("api_url")
    
    if not api_url:
        print("[ERROR] API URL not configured in config.json")
        sys.exit(1)
        
    # Auto-resolve Monthly output directory (e.g. D:\workspace-ai\QUYET TOAN 2026\THANG 7\BUYING T7)
    month_folder_name = f"THANG {args.month}"
    buying_folder_name = f"BUYING T{args.month}"
    
    possible_dirs = [
        os.path.join(base_output_dir, month_folder_name, buying_folder_name),
        os.path.join(base_output_dir, month_folder_name),
        base_output_dir
    ]
    
    output_dir = None
    for p in possible_dirs:
        try:
            os.makedirs(p, exist_ok=True)
            output_dir = p
            break
        except Exception:
            continue
            
    if not output_dir:
        output_dir = os.path.join(os.path.expanduser("~"), "QUYET_TOAN_OUTPUT")
        os.makedirs(output_dir, exist_ok=True)
        
    print(f"[INFO] Thư mục lưu file Excel đầu ra: {output_dir}")

    # Fuzzy match company name or use Smart Scan
    company_key = args.company
    tariff_file = company_map.get(company_key)
    
    # Check if mapped file exists
    if tariff_file and not os.path.exists(tariff_file):
        print(f"[WARNING] Mapped tariff file for {company_key} not found: {tariff_file}. Will rescan.")
        tariff_file = None

    if not tariff_file:
        tariff_dir = config.get("tariff_dir")
        if not tariff_dir or not os.path.exists(tariff_dir):
            print(f"[ERROR] Tariff file not mapped for '{company_key}' and tariff_dir is invalid or missing.")
            sys.exit(1)
            
        print(f"[INFO] Smart Scan: Searching for '{company_key}' in {tariff_dir}...")
        found_files = []
        for root, _, files in os.walk(tariff_dir):
            for file in files:
                if (file.endswith(".xls") or file.endswith(".xlsx")) and not file.startswith("~$"):
                    clean_company = remove_accents(company_key.lower())
                    clean_file = remove_accents(file.lower())
                    if clean_company in clean_file:
                        found_files.append(os.path.join(root, file))
        
        if not found_files:
            print(f"[ERROR] Smart Scan failed: No Tariff file found for '{company_key}' in {tariff_dir}.")
            sys.exit(1)
            
        if len(found_files) > 1:
            print(f"[AMBIGUOUS_MATCH] Smart Scan found multiple Tariff files for '{company_key}':")
            for idx, f in enumerate(found_files):
                print(f"  {idx + 1}. {f}")
            print("[ACTION_REQUIRED] Please ask the user to select the correct file, update company_map.json with the chosen path, and re-run.")
            sys.exit(2)
            
        # Exactly one match
        tariff_file = found_files[0]
        print(f"[SUCCESS] Smart Scan found exactly one Tariff file: {tariff_file}")
        
        # Save to map for future use
        company_map[company_key] = tariff_file
        try:
            with open(MAP_FILE, "w", encoding="utf-8") as f:
                json.dump(company_map, f, indent=2, ensure_ascii=False)
            print(f"[INFO] Updated company_map.json with the new path.")
        except Exception as e:
            print(f"[WARNING] Could not save company_map.json: {e}")

    # 1. Fetch Data
    if args.input_file and os.path.exists(args.input_file):
        shipments = fetch_local_excel_data(args.input_file, args.month, args.year, company_key, args.bill_col, args.cont_col, args.tk_col)
        if shipments is None:
            sys.exit(1)
    else:
        api_response = fetch_api_data(api_url, company_key, args.month, args.year)
        if not api_response or not api_response.get("success"):
            error_msg = api_response.get("error") if api_response else "Unknown error"
            print(f"[ERROR] Failed to fetch data: {error_msg}")
            sys.exit(1)
            
        shipments = api_response.get("data", [])
    if not shipments:
        print(f"[INFO] No shipments found for {company_key} in {args.month}/{args.year}.")
        sys.exit(0)
        
    print(f"[INFO] Retrieved {len(shipments)} shipments from API.")

    # 2. Extract Tariff
    tariff_matrix = extract_tariff_matrix(tariff_file)
    if not tariff_matrix:
        print("[WARNING] Tariff extraction failed. Will use default values.")
        
    # 3. Build EFMS Output
    output_rows = []
    current_date = datetime.datetime.now().strftime("%d/%m/%Y")
    
    for sh in shipments:
        so_bill = str(sh.get("so_bill", ""))
        to_khai = str(sh.get("to_khai", "")).strip()
        
        # Skip if no customs declaration (Business Rule for Toàn Lực / Paritas)
        if not to_khai or to_khai.lower() == 'nan':
            continue
            
        phan_luong = sh.get("phan_luong", "")
        cont_40 = sh.get("cont_40", 0)
        total_cont = sh.get("total_cont", 0)
        to_khai_count = sh.get("to_khai_count", 1)
        
        # Determine unit for containers
        cont_unit = "40'DC" if cont_40 > 0 else "20'DC"
        
        # Determine custom clearance price (BO_NOF_OPS)
        nof_price = get_tariff_price(tariff_matrix, phan_luong, total_cont) if tariff_matrix else 380000
        
        # Override for Haohua
        is_haohua_xuat = "haohua" in company_key.lower() and "xuất" in company_key.lower()
        is_haohua_nhap = "haohua" in company_key.lower() and "nhập" in company_key.lower()
        
        if is_haohua_xuat:
            nof_price = 190000
        elif is_haohua_nhap:
            if "xanh" in phan_luong.lower():
                nof_price = 160000
            elif "vàng" in phan_luong.lower() or "vang" in phan_luong.lower():
                nof_price = 190000
            
        charge_codes = [
            {"code": "BO_COF_OPS", "qty": to_khai_count, "unit": "Tờ khai", "price": 22000, "note": ""},
            {"code": "BO_NOF_OPS", "qty": total_cont, "unit": cont_unit, "price": nof_price, "note": ""},
            {"code": "BO_OTH_OPS", "qty": total_cont, "unit": cont_unit, "price": 300000, "note": ""}
        ]
        
        if is_haohua_nhap and ("vàng" in phan_luong.lower() or "vang" in phan_luong.lower()):
            charge_codes.append({
                "code": "BO_OTH_OPS",
                "qty": 1,
                "unit": "shipment",
                "price": 300000,
                "note": "Hiện trường mở tờ khai"
            })
        
        # Add BO_OTH_OPS for phu_thu_hq (Cột O, mapped as dest_hm_20 in API)
        phu_thu_hq = sh.get("dest_hm_20", 0)
        try:
            phu_thu_hq = float(phu_thu_hq) if phu_thu_hq else 0
        except:
            phu_thu_hq = 0
            
        if phu_thu_hq > 0:
            charge_codes.append({
                "code": "BO_OTH_OPS",
                "qty": 1,
                "unit": "shipment",
                "price": phu_thu_hq,
                "note": "phụ thu hải quan"
            })

        # Add BO_OTH_OPS for phu_thu_bx (Cột P, mapped as dest_hm_40 in API)
        phu_thu_bx = sh.get("dest_hm_40", 0)
        try:
            phu_thu_bx = float(phu_thu_bx) if phu_thu_bx else 0
        except:
            phu_thu_bx = 0
            
        if phu_thu_bx > 0:
            charge_codes.append({
                "code": "BO_OTH_OPS",
                "qty": 1,
                "unit": "shipment",
                "price": phu_thu_bx,
                "note": "Phụ thu bốc xếp"
            })
        
        # Remove BO_OTH_OPS for Haohua as requested
        if "haohua" in company_key.lower() and ("xuất" in company_key.lower() or "nhập" in company_key.lower()):
            charge_codes = [c for c in charge_codes if c["code"] != "BO_OTH_OPS" or c.get("note")]
        
        for charge in charge_codes:
            row_data = {
                "is_special": sh.get("is_special", False),
                "raw_bill": sh.get("raw_bill", ""),
                "HBL (*)": so_bill,
                "MBL ": "",
                "Custom No": to_khai,
                "Partner Code (*)": "VDTRUCKINGHD-000",
                "OBH Partner": "",
                "Charge Code (*)": charge["code"],
                "Qty (*)": charge["qty"],
                "Unit (*)": charge["unit"],
                "Unit Price (*)": charge["price"],
                "Currency (*)": "VND",
                "VAT Partner": "",
                "VAT (*)": "",
                "Exchange Date (*)": current_date,
                "Final Exchange (*)": 24410,
                "Invoice No": "00077",
                "Invoice Date": current_date,
                "Serie No": "",
                "Type (*)": "Buying",
                "Note": charge.get("note", "")
            }
            output_rows.append(row_data)
            
    if args.limit:
        output_rows = output_rows[:args.limit]
    df_output = pd.DataFrame(output_rows)
    
    # Filter out special bills to a separate file
    if 'is_special' in df_output.columns:
        df_special = df_output[df_output['is_special'] == True].copy()
        
        # Các file 1-4 không bao gồm các bill có khoảng trắng
        df_normal = df_output[df_output['is_special'] == False].copy()
        
        # Trong file SPECIAL, trả lại nguyên trạng chuỗi gốc (có chứa khoảng trắng)
        if 'raw_bill' in df_special.columns:
            df_special['HBL (*)'] = df_special['raw_bill']
            
        df_special = df_special.drop(columns=['is_special', 'raw_bill'])
        df_normal = df_normal.drop(columns=['is_special', 'raw_bill'])
        
        if not df_special.empty:
            # Sort special file by charge code
            sort_order = {"BO_COF_OPS": 0, "BO_NOF_OPS": 1, "BO_OTH_OPS": 2}
            df_special['sort_key'] = df_special['Charge Code (*)'].map(sort_order)
            df_special['original_idx'] = range(len(df_special))
            df_special = df_special.sort_values(by=['sort_key', 'original_idx']).drop(['sort_key', 'original_idx'], axis=1)
            
            special_path = os.path.join(output_dir, f"LogisticsImport_EFMS_Buying_{company_key}_{args.month:02d}_{args.year}_SPECIAL.xlsx")
            df_special.to_excel(special_path, index=False, engine='openpyxl')
            print(f"[SUCCESS] Saved {len(df_special)} special rows (with spaces) to {special_path}")
            
        df_output = df_normal
    
    # Sort by Note existence first, then Charge Code, but preserve original API HBL order
    df_output['has_note'] = df_output['Note'].apply(lambda x: 1 if pd.notna(x) and str(x).strip() != "" else 0)
    sort_order = {"BO_COF_OPS": 0, "BO_NOF_OPS": 1, "BO_OTH_OPS": 2}
    df_output['sort_key'] = df_output['Charge Code (*)'].map(sort_order)
    df_output['original_idx'] = range(len(df_output))
    df_output = df_output.sort_values(by=['has_note', 'sort_key', 'original_idx']).drop(['has_note', 'sort_key', 'original_idx'], axis=1)
    
    # Save to Excel
    num_splits = max(1, args.split)
    chunk_size = math.ceil(len(df_output) / num_splits)
    
    from collections import Counter
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill
    
    bill_counts = Counter(sh.get("so_bill", "") for sh in shipments)
    duplicate_bills = {bill for bill, count in bill_counts.items() if count > 1}
    
    for i in range(num_splits):
        start_idx = i * chunk_size
        end_idx = min((i + 1) * chunk_size, len(df_output))
        if start_idx >= len(df_output):
            break
            
        df_chunk = df_output.iloc[start_idx:end_idx]
        
        part_suffix = f"_{i+1:02d}" if num_splits > 1 else ""
        output_filename = f"LogisticsImport_EFMS_Buying_{company_key}_{args.month:02d}_{args.year}{part_suffix}.xlsx"
        output_path = os.path.join(output_dir, output_filename)
        
        def save_and_highlight(target_path):
            df_chunk.to_excel(target_path, index=False, engine='openpyxl')
            wb = load_workbook(target_path)
            ws = wb.active
            yellow_fill = PatternFill(start_color='FFFFFF00', end_color='FFFFFF00', fill_type='solid')
            cof_fill = PatternFill(start_color='DDEBF7', end_color='DDEBF7', fill_type='solid')
            nof_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            oth_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
            
            charge_col_idx = None
            for col in range(1, ws.max_column + 1):
                if ws.cell(row=1, column=col).value == "Charge Code (*)":
                    charge_col_idx = col
                    break
                    
            for row in range(2, ws.max_row + 1):
                hbl_val = str(ws.cell(row=row, column=1).value)
                charge_val = str(ws.cell(row=row, column=charge_col_idx).value) if charge_col_idx else ""
                
                fill = None
                if charge_val == "BO_COF_OPS":
                    fill = cof_fill
                elif charge_val == "BO_NOF_OPS":
                    fill = nof_fill
                elif charge_val == "BO_OTH_OPS":
                    fill = oth_fill
                    
                if fill:
                    for col in range(1, ws.max_column + 1):
                        ws.cell(row=row, column=col).fill = fill
                
                if duplicate_bills and hbl_val in duplicate_bills:
                    ws.cell(row=row, column=1).fill = yellow_fill
                    
            wb.save(target_path)

        try:
            save_and_highlight(output_path)
            print(f"[SUCCESS] Successfully generated EFMS Buying file: {output_path}")
            print(f"[INFO] File {i+1} rows: {len(df_chunk)}")
        except Exception as e:
            # Fallback 1: Try saving with timestamp if file is open in Excel or locked
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            alt_filename = f"LogisticsImport_EFMS_Buying_{company_key}_{args.month:02d}_{args.year}{part_suffix}_{ts}.xlsx"
            alt_path = os.path.join(output_dir, alt_filename)
            try:
                save_and_highlight(alt_path)
                print(f"[SUCCESS] Saved with timestamp due to file lock: {alt_path}")
            except Exception as e2:
                # Fallback 2: Local user directory
                local_dir = os.path.join(os.path.expanduser("~"), "QUYET_TOAN_OUTPUT")
                os.makedirs(local_dir, exist_ok=True)
                local_path = os.path.join(local_dir, output_filename)
                try:
                    save_and_highlight(local_path)
                    print(f"[SUCCESS] Saved to user fallback directory: {local_path}")
                except Exception as e3:
                    print(f"[ERROR] Failed to save Excel file: {e3}")
            
    if duplicate_bills:
        print(f"[INFO] Highlighted {len(duplicate_bills)} duplicate bills in yellow.")
    print(f"[INFO] Total generated rows across {num_splits} file(s): {len(df_output)}")

if __name__ == "__main__":
    main()
