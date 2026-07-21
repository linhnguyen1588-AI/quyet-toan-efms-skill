import asyncio
import os
import sys
import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
from playwright.async_api import async_playwright

sys.stdout.reconfigure(encoding='utf-8')

import json

async def main():
    if len(sys.argv) > 1 and sys.argv[1].strip():
        excel_path = sys.argv[1].strip()
    else:
        default_catlai = r"D:\workspace-ai\HOA DON EPORT CAT LAI\HOA DON EPORT CAT LAI 01.07.xlsx"
        default_haohua = r"D:\workspace-ai\HOA DON\HAOHUA\HAOHUA XUAT EPORT.xlsx"
        if os.path.exists(default_catlai):
            excel_path = default_catlai
        elif os.path.exists(default_haohua):
            excel_path = default_haohua
        else:
            excel_path = default_catlai
        
    out_dir = os.path.join(os.path.dirname(excel_path), "HOA DON")
    if not os.path.exists(out_dir):
        os.makedirs(out_dir, exist_ok=True)
        
    print(f"Đang đọc file Excel từ: {excel_path}")
    try:
        df = pd.read_excel(excel_path)
        booking_col = None
        for col in df.columns:
            col_str = str(col).strip().upper()
            if 'BOOK' in col_str or 'BILL' in col_str or 'CONT' in col_str:
                booking_col = col
                break
        if booking_col:
            print(f"-> Đã phát hiện cột chứa mã: '{booking_col}'")
            bookings = df[booking_col].dropna().astype(str).unique().tolist()
        else:
            print("-> Dùng cột đầu tiên trong file Excel...")
            bookings = df.iloc[:, 0].dropna().astype(str).unique().tolist()
            
        bookings = [b.strip() for b in bookings if str(b).strip() and str(b).strip().lower() != 'nan']
        target_books = bookings
    except Exception as e:
        print("Lỗi đọc file Excel:", e)
        return

    if not target_books:
        print("Không tìm thấy số Book nào trong file!")
        return

    print(f"Sẽ tiến hành quét {len(target_books)} Book: {target_books}")
    failed_books = []

    async with async_playwright() as p:
        # Headless background mode: runs silently without opening visible browser window
        browser = await p.chromium.launch(headless=True)
        context = await browser.new_context(accept_downloads=True)
        page = await context.new_page()
        
        print("\n==================================================")
        print("SIÊU ROBOT API CHẠY NGẦM ĐÃ KÍCH HOẠT (HEADLESS MODE)!")
        print("Tự động đăng nhập ePort Tân Cảng Cát Lái & tải hóa đơn PDF...")
        print("==================================================\n")
        
        await page.goto("https://eport.saigonnewport.com.vn/")
        
        # Tự động đăng nhập ngầm tài khoản ePort
        try:
            await page.wait_for_timeout(2000)
            tax_loc = page.locator("input[type='text']:visible").first
            await tax_loc.fill("0314436809")
            pass_loc = page.locator("input[type='password']:visible").first
            await pass_loc.fill("Sotrans1234@")
            login_btn = page.locator("button:visible, input[type='submit']:visible").filter(has_text="Đăng nhập").first
            await login_btn.click()
            print("[INFO] Đã gửi thông tin đăng nhập tự động vào ePort Tân Cảng...")
            await page.wait_for_timeout(3000)
        except Exception as e:
            print(f"[WARNING] Đăng nhập tự động: {e}")

        await page.goto("https://eport.saigonnewport.com.vn/FullContainerDelivery")
        await page.wait_for_timeout(2000)

        print("\n[+] Đã hoàn tất đăng nhập ngầm! Bắt đầu càn quét qua API...")
        api_context = context.request
        
        # 1. Gọi API Search để lấy toàn bộ BatchNo trong 90 ngày (chia làm 3 chặng 30 ngày)
        print("  -> Đang nạp danh sách Booking từ máy chủ ePort...")
        order_map = {}
        for m in range(3):
            date_to_dt = datetime.now() - timedelta(days=m*30)
            date_from_dt = date_to_dt - timedelta(days=29) # 29 days to be strictly < 30
            date_to = date_to_dt.strftime("%Y-%m-%dT%H:%M:%S.000Z")
            date_from = date_from_dt.strftime("%Y-%m-%dT%H:%M:%S.000Z")
            
            for oper_type in ["", "FRECV", "HBCX", "FDELE", "FDISP", "EDO"]:
                search_payload = {
                    "DateFrom": date_from,
                    "DateTo": date_to,
                    "OperType": oper_type,
                    "HasEDO": "False"
                }
                search_resp = await api_context.post(
                    "https://eport.saigonnewport.com.vn/BatchNoList/SeachBatchNoList", 
                    data=json.dumps(search_payload),
                    headers={
                        "Content-Type": "application/json; charset=UTF-8",
                        "X-Requested-With": "XMLHttpRequest",
                        "Referer": "https://eport.saigonnewport.com.vn/FullContainerDelivery"
                    }
                )
                try:
                    search_data = await search_resp.json()
                except Exception as e:
                    print(f"  [!] Lỗi JSON: {e}, Status: {search_resp.status}")
                    print(f"  [!] Response body: {(await search_resp.text())[:500]}")
                    continue
                    
                data_arr = search_data.get("Data")
                if not data_arr:
                    print(f"  [!] API trả về không có Data: {str(search_data)[:200]}")
                    continue
                    
                for item in data_arr:
                    notes = str(item.get("NOTES", "")).strip()
                    order_id = item.get("ORDER_ID")
                    if notes and order_id:
                        for book in target_books:
                            if book in notes:
                                order_map[book] = order_id
        
        print(f"  [v] Đã tra cứu thành công {len(order_map)}/{len(target_books)} Book có trên hệ thống!")
        
        # 2. Xử lý từng book
        for book in target_books:
            print(f"\n>>> Đang xử lý ngầm Book: {book}")
            if book not in order_map:
                print(f"  [!] Book {book} không tồn tại hoặc đã quá hạn trên ePort.")
                failed_books.append(book)
                continue
                
            order_id = order_map[book]
            try:
                # Gọi API Init hóa đơn
                init_resp = await api_context.post(
                    "https://eport.saigonnewport.com.vn/InvoiceEcom/Init", 
                    data=json.dumps({"orderId": order_id}),
                    headers={"Content-Type": "application/json; charset=UTF-8"}
                )
                init_data = await init_resp.json()
                
                invoices = init_data.get("Data", {}).get("Item2", [])
                if not invoices:
                    print(f"  [!] Book {book} chưa có hóa đơn nào được xuất.")
                    failed_books.append(book)
                    continue
                    
                print(f"  -> Tìm thấy {len(invoices)} hóa đơn. Đang tải...")
                for idx, inv in enumerate(invoices):
                    fkey = inv.get("INVOICE_FKEY")
                    if not fkey: continue
                    
                    # Gọi API Download POST để lấy token
                    await api_context.post(
                        "https://eport.saigonnewport.com.vn/Payment/DownLoadInvoice_POST", 
                        data=json.dumps({"fKey": fkey}),
                        headers={"Content-Type": "application/json; charset=UTF-8"}
                    )
                    
                    # Gọi API Download GET để lấy file PDF
                    pdf_resp = await api_context.get("https://eport.saigonnewport.com.vn/Payment/DownLoadInvoice_PDF_GET")
                    
                    if pdf_resp.status == 200:
                        suffix = f"_{idx + 1}"
                        file_path = os.path.join(out_dir, f"{book}{suffix}.pdf")
                            
                        with open(file_path, "wb") as f:
                            f.write(await pdf_resp.body())
                        print(f"  [v] TẢI XONG: {file_path}")
                    else:
                        print(f"  [!] Lỗi khi tải PDF cho Book {book}, HTTP {pdf_resp.status}")
                        
            except Exception as e:
                print(f"  [!] Lỗi khi gọi API cho Book {book}: {e}")
                failed_books.append(book)
                
        print("\n🎉 HOÀN THÀNH TẢI QUA API SIÊU TỐC!")
        await asyncio.sleep(2)
        await browser.close()
        
    # Xử lý tô màu vàng cho các book lỗi
    if failed_books:
        print(f"\n[+] Đang tô vàng {len(failed_books)} Book lỗi...")
        try:
            wb = openpyxl.load_workbook(excel_path)
            ws = wb.active
            
            booking_col_idx = None
            for row in ws.iter_rows(min_row=1, max_row=5):
                for cell in row:
                    if cell.value and "BOOKING" in str(cell.value).upper():
                        booking_col_idx = cell.column
                        break
                if booking_col_idx: break
            
            if booking_col_idx:
                yellow_fill = PatternFill(start_color="FFFFFF00", end_color="FFFFFF00", fill_type="solid")
                for row in ws.iter_rows(min_row=2):
                    cell_val = str(row[booking_col_idx - 1].value).strip() if row[booking_col_idx - 1].value else ""
                    if cell_val in failed_books:
                        for cell in row:
                            cell.fill = yellow_fill
                try:
                    wb.save(excel_path)
                    print(f"  [v] Đã tô vàng xong file Excel!")
                except PermissionError:
                    print(f"  [!] LỖI: Vui lòng tắt file Excel rồi chạy lại để tô màu.")
        except Exception as e:
            pass

if __name__ == "__main__":
    asyncio.run(main())
